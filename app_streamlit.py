"""
SKAB CONTRÔLE INTERNE - APPLICATION WEB DE CONSOLIDATION
Streamlit App pour gérer la consolidation des données de contrôle

Installation:
    pip install streamlit pandas openpyxl

Lancement:
    streamlit run app.py
"""

import streamlit as st
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import io
import zipfile

# ========== CONFIG ==========
st.set_page_config(
    page_title="SKAB - Consolidation CI",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Styles CSS personnalisés
st.markdown("""
    <style>
    /* Thème principal */
    :root {
        --primary: #1E40AF;
        --secondary: #059669;
        --danger: #DC2626;
        --warning: #F59E0B;
        --light: #F3F4F6;
        --dark: #111827;
    }
    
    /* Font personnalisée */
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;700&family=JetBrains+Mono:wght@400;500&display=swap');
    
    * {
        font-family: 'Poppins', sans-serif;
    }
    
    code {
        font-family: 'JetBrains Mono', monospace;
    }
    
    /* En-têtes */
    h1, h2, h3 {
        color: #1E40AF;
        font-weight: 700;
    }
    
    /* Boutons */
    .stButton > button {
        background: linear-gradient(135deg, #1E40AF 0%, #059669 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 10px 24px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 20px rgba(30, 64, 175, 0.3);
    }
    
    /* Conteneurs */
    .metric-card {
        background: linear-gradient(135deg, #F0F9FF 0%, #ECFDF5 100%);
        padding: 20px;
        border-radius: 12px;
        border-left: 4px solid #1E40AF;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
    }
    
    /* Alerts */
    .alert-success {
        background: #ECFDF5;
        border-left: 4px solid #059669;
        padding: 15px;
        border-radius: 8px;
        color: #065F46;
    }
    
    .alert-danger {
        background: #FEF2F2;
        border-left: 4px solid #DC2626;
        padding: 15px;
        border-radius: 8px;
        color: #7F1D1D;
    }
    
    .alert-warning {
        background: #FFFBEB;
        border-left: 4px solid #F59E0B;
        padding: 15px;
        border-radius: 8px;
        color: #78350F;
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%);
    }
    </style>
""", unsafe_allow_html=True)

# ========== FONCTIONS ==========

def extraire_code_controleur(nom_fichier):
    """Extrait le code contrôleur du nom de fichier"""
    try:
        parties = nom_fichier.replace('.xlsx', '').split('_')
        return f"{parties[0]}_{parties[1]}"
    except:
        return "INCONNU"

def extraire_nom_controleur(nom_fichier):
    """Extrait le nom du contrôleur"""
    try:
        parties = nom_fichier.replace('.xlsx', '').split('_')
        if len(parties) >= 3:
            return ' '.join(parties[2:]).upper()
    except:
        pass
    return "INCONNU"

def lire_donnees_controleur(fichier, feuille_source, index_upload):
    """Lit les données d'un fichier contrôleur"""
    try:
        # Lire sans spécifier l'en-tête pour trouver les vraies données
        df = pd.read_excel(fichier, sheet_name=feuille_source, header=None)
        
        # Trouver la ligne d'en-tête
        header_row = None
        for idx, row in df.iterrows():
            if any(isinstance(cell, str) and len(str(cell).strip()) > 2 for cell in row):
                header_row = idx
                break
        
        if header_row is not None:
            # Relire avec le bon en-tête
            df = pd.read_excel(fichier, sheet_name=feuille_source, header=header_row)
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            
            return df, True
    except Exception as e:
        st.error(f"Erreur lecture {feuille_source}: {str(e)}")
    
    return pd.DataFrame(), False

def consolider_donnees(fichiers_upload, feuille_source):
    """Consolide les données de plusieurs fichiers"""
    donnees_consolidees = []
    rapport = []
    
    for idx, fichier_upload in enumerate(fichiers_upload):
        nom_fichier = fichier_upload.name
        code_ctrl = extraire_code_controleur(nom_fichier)
        nom_ctrl = extraire_nom_controleur(nom_fichier)
        
        df, succes = lire_donnees_controleur(fichier_upload, feuille_source, idx)
        
        if succes and len(df) > 0:
            # Ajouter métadonnées
            df['Contrôleur'] = nom_ctrl
            df['Code_Contrôleur'] = code_ctrl
            df['Fichier_source'] = nom_fichier
            df['Date_import'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            
            donnees_consolidees.append(df)
            rapport.append({
                'fichier': nom_fichier,
                'controleur': nom_ctrl,
                'lignes': len(df),
                'statut': '✅ OK'
            })
        else:
            rapport.append({
                'fichier': nom_fichier,
                'controleur': nom_ctrl,
                'lignes': 0,
                'statut': '⚠️ Aucune donnée'
            })
    
    if donnees_consolidees:
        result = pd.concat(donnees_consolidees, ignore_index=True, sort=False)
        return result, rapport
    else:
        return pd.DataFrame(), rapport

def generer_fichier_maitre(consolidations_dict, fichier_template):
    """Génère un fichier MAÎTRE avec les données consolidées"""
    try:
        # Charger le template
        wb = load_workbook(fichier_template)
        
        # Écrire les consolidations
        for feuille_maitre, df in consolidations_dict.items():
            if feuille_maitre in wb.sheetnames:
                ws = wb[feuille_maitre]
                
                # Effacer les anciennes données (à partir de la row 5)
                for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None
                
                # Écrire les nouvelles données
                for r_idx, row in enumerate(df.values, 5):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Sauvegarder en mémoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, True
    except Exception as e:
        st.error(f"Erreur génération fichier: {str(e)}")
        return None, False

# ========== INTERFACE STREAMLIT ==========

# En-tête
col1, col2 = st.columns([0.8, 0.2])
with col1:
    st.markdown("# 📊 SKAB - Consolidation Contrôle Interne")
    st.markdown("### Système de consolidation automatisée 2026")
with col2:
    st.metric("État", "🟢 Actif", "+1")

st.divider()

# Sidebar - Navigation
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    
    mode = st.radio(
        "Mode d'utilisation:",
        ["📤 Importer fichiers", "🎯 Consolider", "📥 Télécharger", "📖 Guide"],
        key="mode_nav"
    )
    
    st.divider()
    
    st.markdown("## 📋 Informations")
    st.info("""
    **Application SKAB v1.0**
    
    Gestion consolidée des missions de contrôle interne pour le Groupe SKAB.
    
    **Contacts:**
    - DAF: DIGNOU Élie
    - CD CI: Francis NKIMI NGASSAM
    """)

# ========== MODE 1: IMPORTER FICHIERS ==========
if mode == "📤 Importer fichiers":
    st.subheader("📤 Importer les fichiers contrôleurs")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 1. Sélectionnez les fichiers contrôleurs")
        fichiers = st.file_uploader(
            "Uploadez les fichiers contrôleurs (CI-XXX_NOM_MOIS.xlsx)",
            type="xlsx",
            accept_multiple_files=True,
            key="file_upload"
        )
        
        if fichiers:
            st.success(f"✅ {len(fichiers)} fichier(s) chargé(s)")
            
            # Afficher les fichiers chargés
            st.markdown("**Fichiers détectés:**")
            for f in fichiers:
                code = extraire_code_controleur(f.name)
                nom = extraire_nom_controleur(f.name)
                st.text(f"  • {code} - {nom}")
    
    with col2:
        st.markdown("### 2. Configurez la consolidation")
        
        feuilles_dispo = ["MES_MISSIONS", "POINTS_CONTROLE", "ANOMALIES", "PLANS_ACTION"]
        feuilles_select = st.multiselect(
            "Feuilles à consolider:",
            feuilles_dispo,
            default=feuilles_dispo,
            key="feuilles_select"
        )
        
        st.markdown("### 3. Options")
        inclure_metadata = st.checkbox("Inclure métadonnées (Contrôleur, Date, Source)", value=True)
        supprimer_doublons = st.checkbox("Supprimer les doublons", value=False)
    
    st.divider()
    
    # Stockage en session state
    if 'donnees_consolidees' not in st.session_state:
        st.session_state.donnees_consolidees = {}
    
    # Consolider
    if st.button("🚀 Lancer la consolidation", key="btn_consolider", use_container_width=True):
        if fichiers and feuilles_select:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for idx, feuille in enumerate(feuilles_select):
                status_text.text(f"Consolidation de {feuille}...")
                
                df_consolidated, rapport = consolider_donnees(fichiers, feuille)
                st.session_state.donnees_consolidees[feuille] = {
                    'data': df_consolidated,
                    'rapport': rapport
                }
                
                progress_bar.progress((idx + 1) / len(feuilles_select))
            
            status_text.text("✅ Consolidation terminée!")
            st.success("Les données sont prêtes pour le téléchargement")
        else:
            st.error("❌ Veuillez sélectionner des fichiers et des feuilles")

# ========== MODE 2: CONSOLIDER ==========
elif mode == "🎯 Consolider":
    st.subheader("🎯 Aperçu de la consolidation")
    
    if st.session_state.donnees_consolidees:
        
        # Statistiques globales
        col1, col2, col3, col4 = st.columns(4)
        
        total_lignes = sum(len(data['data']) for data in st.session_state.donnees_consolidees.values())
        total_fichiers = len([r for data in st.session_state.donnees_consolidees.values() for r in data['rapport']])
        taux_success = sum(1 for data in st.session_state.donnees_consolidees.values() 
                          for r in data['rapport'] if '✅' in r['statut']) / max(total_fichiers, 1) * 100
        
        with col1:
            st.metric("📊 Lignes consolidées", total_lignes)
        with col2:
            st.metric("📁 Fichiers traités", total_fichiers)
        with col3:
            st.metric("✅ Taux de succès", f"{taux_success:.0f}%")
        with col4:
            st.metric("🕐 Date/Heure", datetime.now().strftime("%H:%M:%S"))
        
        st.divider()
        
        # Détail par feuille
        tabs = st.tabs([f"📥 {feuille}" for feuille in st.session_state.donnees_consolidees.keys()])
        
        for tab, (feuille, data) in zip(tabs, st.session_state.donnees_consolidees.items()):
            with tab:
                col1, col2 = st.columns([0.6, 0.4])
                
                with col1:
                    st.markdown(f"### {feuille}")
                    st.markdown(f"**{len(data['data'])} lignes** consolidées")
                    
                    if len(data['data']) > 0:
                        st.dataframe(data['data'].head(10), use_container_width=True, height=300)
                        
                        if st.checkbox(f"Voir tous les détails ({feuille})", key=f"show_all_{feuille}"):
                            st.dataframe(data['data'], use_container_width=True)
                
                with col2:
                    st.markdown("### 📋 Rapport de consolidation")
                    df_rapport = pd.DataFrame(data['rapport'])
                    
                    # Colorer le statut
                    def style_statut(val):
                        if '✅' in val:
                            return 'background-color: #ECFDF5; color: #065F46;'
                        elif '⚠️' in val:
                            return 'background-color: #FFFBEB; color: #78350F;'
                        else:
                            return 'background-color: #FEF2F2; color: #7F1D1D;'
                    
                    st.dataframe(df_rapport, use_container_width=True, hide_index=True)
    else:
        st.info("💡 Veuillez d'abord importer et consolider vos fichiers")

# ========== MODE 3: TÉLÉCHARGER ==========
elif mode == "📥 Télécharger":
    st.subheader("📥 Télécharger les résultats")
    
    if st.session_state.donnees_consolidees:
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📊 Export Excel")
            
            # Template MAÎTRE
            fichier_template = st.file_uploader(
                "Chargez votre fichier MAÎTRE (optionnel)",
                type="xlsx",
                key="template_upload"
            )
            
            if st.button("💾 Générer fichier MAÎTRE", use_container_width=True):
                if fichier_template:
                    # Convertir les DataFrames
                    consolidations_dict = {
                        feuille: data['data'] 
                        for feuille, data in st.session_state.donnees_consolidees.items()
                    }
                    
                    output, succes = generer_fichier_maitre(consolidations_dict, fichier_template)
                    
                    if succes:
                        st.download_button(
                            label="⬇️ Télécharger fichier MAÎTRE",
                            data=output,
                            file_name=f"SKAB_Maitre_Consolidation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_download_maitre"
                        )
                else:
                    st.warning("⚠️ Veuillez charger le fichier MAÎTRE")
        
        with col2:
            st.markdown("### 📑 Export par feuille")
            
            feuille_select = st.selectbox(
                "Sélectionnez une feuille:",
                st.session_state.donnees_consolidees.keys(),
                key="feuille_export"
            )
            
            if feuille_select:
                df_export = st.session_state.donnees_consolidees[feuille_select]['data']
                
                # Export CSV
                csv = df_export.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 Télécharger en CSV",
                    data=csv,
                    file_name=f"{feuille_select}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key=f"btn_csv_{feuille_select}"
                )
                
                # Export Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_export.to_excel(writer, sheet_name=feuille_select, index=False)
                output.seek(0)
                
                st.download_button(
                    label="📊 Télécharger en Excel",
                    data=output,
                    file_name=f"{feuille_select}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"btn_excel_{feuille_select}"
                )
    else:
        st.info("💡 Veuillez d'abord consolider vos fichiers")

# ========== MODE 4: GUIDE ==========
elif mode == "📖 Guide":
    st.subheader("📖 Guide d'utilisation")
    
    with st.expander("🎯 Étape 1 : Préparer les fichiers contrôleurs", expanded=True):
        st.markdown("""
        1. **Dupliquez le template** pour chaque contrôleur
        2. **Renommez chaque fichier** avec le format: `CI_XXX_NOM_MOIS.xlsx`
           - Exemple: `CI_001_NKIMI_Mai2026.xlsx`
        3. **Remplissez les paramètres** du contrôleur
        4. **Saisissez les données** dans:
           - MES_MISSIONS
           - POINTS_CONTROLE
           - ANOMALIES
           - PLANS_ACTION
        """)
    
    with st.expander("📤 Étape 2 : Importer les fichiers"):
        st.markdown("""
        1. Cliquez sur **"Importer fichiers"**
        2. Sélectionnez tous les fichiers contrôleurs
        3. Choisissez les feuilles à consolider
        4. Cliquez sur **"Lancer la consolidation"**
        """)
    
    with st.expander("🎯 Étape 3 : Vérifier les données"):
        st.markdown("""
        1. Cliquez sur **"Consolider"**
        2. Vérifiez les statistiques et le rapport
        3. Visualisez un aperçu des données
        4. Assurez-vous que tous les fichiers sont correctement traités
        """)
    
    with st.expander("📥 Étape 4 : Télécharger les résultats"):
        st.markdown("""
        1. Cliquez sur **"Télécharger"**
        2. **Option A**: Charger votre fichier MAÎTRE et générer la version consolidée
        3. **Option B**: Exporter chaque feuille en CSV ou Excel
        """)
    
    st.divider()
    
    with st.expander("❓ Foire aux questions"):
        st.markdown("""
        ### Quel format de nom de fichier ?
        `CI_XXX_NOM_MOIS.xlsx` - Exemple: `CI_001_NKIMI_Mai2026.xlsx`
        
        ### Que faire en cas d'erreur ?
        - Vérifiez le nom du fichier
        - Vérifiez que la feuille existe dans le fichier
        - Assurez-vous que les données ne sont pas vides
        
        ### Puis-je consolider des fichiers d'années différentes ?
        Oui, les données seront consolidées ensemble
        
        ### Comment supprimer les doublons ?
        Activez l'option "Supprimer les doublons" lors de l'import
        
        ### Puis-je réutiliser les données consolidées ?
        Oui, téléchargez en Excel ou CSV et réutilisez
        """)

# Footer
st.divider()
st.markdown("""
<div style='text-align: center; color: #666; font-size: 0.9em;'>
    <p>📊 SKAB Contrôle Interne 2026 | Groupe SKAB | Powered by Streamlit</p>
    <p>DAF: DIGNOU Élie | CD CI: Francis NKIMI NGASSAM</p>
</div>
""", unsafe_allow_html=True)
